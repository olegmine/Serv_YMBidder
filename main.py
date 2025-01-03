import asyncio
import re
import math
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import signal
import aiohttp
import json
from typing import Optional, Dict, Any, Set
from scr.logger import logger
from scr.data_writer import write_sheet_data
from scr.data_fetcher import get_sheet_data
from scr.yandex_market_report import get_yandex_market_report
from scr.update_data_ym import compare_prices_and_create_for_update, update_dataframe, first_write_df
from scr.update_ym import update_price_ym

# Глобальные переменные для управления состоянием
is_running = True
active_tasks: Set[asyncio.Task] = set()
shutdown_event = asyncio.Event()
DEBUG = True

class MarketplaceConfig:
    def __init__(self,
                 user_id: str,
                 sample_spreadsheet_id: str,
                 update_interval_minutes: int,
                 api_yandex_market: Optional[str] = None,
                 business_id_yandex_market: Optional[str] = None,
                 yandex_market_range: Optional[str] = None,
                 market_name: Optional[str] = None,
                 user_email: Optional[str] = None,
                 phone_number: Optional[str] = None,
                 price_decrease_lower: Optional[int] = None,
                 price_decrease_upper: Optional[int] = None,
                 market_white_list: Optional[str] = None):  # Добавлен новый параметр
        self.user_id = user_id
        self.sample_spreadsheet_id = sample_spreadsheet_id
        self.update_interval_minutes = int(update_interval_minutes)
        self.api_yandex_market = api_yandex_market
        self.business_id_yandex_market = business_id_yandex_market
        self.yandex_market_range = yandex_market_range
        self.market_name = market_name
        self.user_email = user_email
        self.phone_number = phone_number
        self.price_decrease_lower = int(price_decrease_lower) if price_decrease_lower is not None else None
        self.price_decrease_upper = int(price_decrease_upper) if price_decrease_upper is not None else None
        # Преобразование строки с разделителями-запятыми в список
        self.market_white_list = [x.strip() for x in market_white_list.split(',')] if market_white_list else []
    def get_user_info(self) -> str:
        """Возвращает информацию о пользователе для логов"""
        return f"[ID: {self.user_id}, Email: {self.user_email}, Тел: {self.phone_number}]"

    def has_user_config(self) -> bool:
        return all([self.api_yandex_market, self.business_id_yandex_market,self.yandex_market_range])

async def save_debug_csv(df: pd.DataFrame, filename: str) -> None:
    if DEBUG:
        try:
            await asyncio.to_thread(df.to_csv, filename, index=False)
            logger.debug(f"Сохранен отладочный CSV: {filename}")
        except IOError as e:
            logger.error(f"Ошибка при сохранении отладочного CSV {filename}: {str(e)}")


async def process_yandex_market_data(session: aiohttp.ClientSession, config: MarketplaceConfig) -> Dict[str, Any]:
    """Обработка данных Яндекс.Маркета"""
    ym_logger = logger.bind(marketplace="YandexMarket")
    ym_flag = None
    sheets_flag = None

    # Валидация входных данных
    if not all([config.sample_spreadsheet_id, config.market_name, config.yandex_market_range,
                config.business_id_yandex_market, config.api_yandex_market]):
        raise ValueError("Отсутствуют необходимые параметры конфигурации")

    try:
        if config.market_white_list :
            my_market = config.market_white_list
        else:
            my_market = ['sample']
        # Инициализация параметров
        market_config = {
            'spreadsheet_id': config.sample_spreadsheet_id,
            'market_name': config.market_name,
            'range_name': config.market_name,
            'sheet_range': config.yandex_market_range,
            'api_key': config.api_yandex_market,
            'business_id': config.business_id_yandex_market,
            'safe_user_name': re.sub(r'[^\w\-_]', '_', config.user_id),
            'safe_market_name': re.sub(r'[^\w\-_]', '_', config.market_name)

        }

        COLUMNS_FULL = {
            'seller_id': 'SHOP_SKU',
            'name': 'OFFER',
            'link': 'LINK',
            'price': 'MERCH_PRICE_WITH_PROMOS',
            'stop': 'STOP',
            'mp_on_market': 'PRICE_VALUE_ON_MARKET',
            'market_with_mp': 'SHOP_WITH_BEST_PRICE_ON_MARKET',
            'prim': 'PRIM'
        }

        COLUMNS_TO_KEEP = [
            'SHOP_SKU', 'OFFER','MERCH_PRICE_WITH_PROMOS',
            'PRICE_GREEN_THRESHOLD', 'PRICE_RED_THRESHOLD',
            'SHOP_WITH_BEST_PRICE_ON_MARKET', 'PRICE_VALUE_ON_MARKET'
        ]
        SQLITE_DB_NAME = f"databases/{market_config['safe_user_name']}_data_{market_config['safe_market_name']}.db"
        try:
            # Получение данных из Google Sheets
            ym_logger.info(f"Получение данных из Google Sheets для {market_config['range_name']}")
            df_from_sheets = await get_sheet_data(market_config['spreadsheet_id'], market_config['sheet_range'])
            sheets_flag = True
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            await save_debug_csv(df_from_sheets, f"report/{market_config['range_name']}{current_time}_first.csv")
        except Exception as e:
            sheets_flag = False
            ym_logger.error(f"Не удалось получить данные из Гугл таблиц для пользователя {config.user_id} "
                            f"с email {config.user_email}")


        # Получение отчета с Яндекс.Маркета
        try:
            ym_report_df = await get_yandex_market_report(market_config['api_key'], market_config['business_id'])
            # Проверяем наличие колонок и логируем отсутствующие
            missing_columns = [col for col in COLUMNS_TO_KEEP if col not in ym_report_df.columns]
            if missing_columns:
                logger.warning(
                    f"Следующие колонки отсутствуют в полученном отчете: {', '.join(missing_columns)}",
                    extra={
                        'market_name': config.market_name,
                        'user_name': config.user_id
                    }
                )

            # Фильтруем датафрейм, оставляя только те колонки, которые есть и в COLUMNS_TO_KEEP, и в датафрейме
            available_columns = [col for col in COLUMNS_TO_KEEP if col in ym_report_df.columns]
            ym_report_df = ym_report_df[available_columns]

            # Проверяем наличие колонки 'PRICE_VALUE_ON_MARKET' перед выполнением dropna
            if 'PRICE_VALUE_ON_MARKET' in ym_report_df.columns:
                ym_report_df = await asyncio.to_thread(
                    lambda: ym_report_df.dropna(subset=['PRICE_VALUE_ON_MARKET'])
                )
            else:
                logger.warning(
                    "Колонка 'PRICE_VALUE_ON_MARKET' отсутствует в отчете, пропускаем удаление NA значений",
                    extra={
                        'market_name': config.market_name,
                        'user_name': config.user_id
                    }
                )
            await save_debug_csv(ym_report_df, f"report/{market_config['range_name']}{current_time}_ym_report.csv")
            ym_flag = True
        except Exception as e:
            ym_flag = False
            ym_logger.error(f"Ошибка при получении отчета с Яндекс.Маркета: {str(e)}")
            raise

        if ym_flag == True and sheets_flag == False:
            df_for_write = await first_write_df(ym_report_df)
            ym_logger.warning(
                f"Таблица Google пуста для клиента {config.user_id}, записываю данные из личного кабинета"
            )
            # Создаем копию DataFrame
            df_to_write = df_for_write.copy()
            # Получаем названия колонок
            column_names = pd.DataFrame([df_to_write.columns.tolist()], columns=df_to_write.columns)
            # Сначала добавляем названия колонок, затем данные
            df_to_write = pd.concat([column_names, df_to_write], axis=0, ignore_index=True)

            await write_sheet_data(
                df_to_write,
                config.sample_spreadsheet_id,
                config.yandex_market_range
            )

        if ym_flag == True and sheets_flag == True:
            # Обновление и сравнение данных
            try:
                df_from_sheets = df_from_sheets.iloc[1:].copy()

                updated_df = await update_dataframe(df1=df_from_sheets,
                                                    df2=ym_report_df,
                                                    column_names=COLUMNS_FULL,
                                                    user_name=config.user_id,
                                                    market_name=config.market_name)
                updated_df.to_csv(f"{config.market_name}.csv")
            except Exception as e:
                ym_logger.error(f"Ошибка при обновлении данных: {str(e)}")
                raise
            try:
                updated_df_final, for_update_df = await compare_prices_and_create_for_update(
                    df=updated_df,
                    column_names =COLUMNS_FULL,
                    my_market=my_market,
                    db_file=SQLITE_DB_NAME,
                    username=config.user_id,
                    marketname=config.market_name,
                    min_price_diff=config.price_decrease_lower,
                    max_price_diff=config.price_decrease_upper
                )

                await write_sheet_data(
                    updated_df_final,
                    market_config['spreadsheet_id'],
                    market_config['sheet_range'].replace('1', '3')
                )
            except Exception as e:
                ym_logger.error(f"Ошибка при сравнении данных: {str(e)}")
                raise

            # Обновление цен через API
            if not for_update_df.empty:
                ym_logger.warning(f"Обновление цен через API для {market_config['range_name']}", importance="high")
                try:
                    await update_price_ym(
                        for_update_df,
                        market_config['api_key'],
                        market_config['business_id'],
                        "SHOP_SKU",
                        "MERCH_PRICE_WITH_PROMOS",
                        'discount_base',
                        debug=DEBUG,
                        marketname=config.market_name,
                        username=config.user_id
                    )
                    ym_logger.warning("Завершено обновление цен через API")
                except Exception as e:
                    ym_logger.error(f"Ошибка при обновлении цен через API: {str(e)}")
                    raise


            try:
                await save_debug_csv(updated_df, f"report_ym/{market_config['range_name']}{current_time}_updated.csv")
                await save_debug_csv(for_update_df, f"report_ym/{market_config['range_name']}{current_time}_for_update.csv")
            except Exception as e:
                ym_logger.info('Не удалось сохранить один из датафреймов')

            ym_logger.info(f"Обработка завершена успешно для {market_config['range_name']}")
            return {
                'status': 'success',
                'marketplace': 'YandexMarket',
                'rows_processed': len(df_from_sheets),
                'rows_updated': len(for_update_df)
            }

    except Exception as e:
        error_details = {
            'user_id': config.user_id,
            'market_name': config.market_name,
            'range': config.yandex_market_range,
            'error': str(e)
        }
        ym_logger.error("Критическая ошибка при обновлении данных Яндекс.Маркета", **error_details)

        return {
            'status': 'error',
            'marketplace': 'YandexMarket',
            'error': str(e),
            'details': error_details
        }


def get_users_config_from_excel(filename: str) -> list:
    """Читает конфигурацию пользователей из Excel файла"""
    try:
        wb = load_workbook(filename=filename, read_only=True)
        sheet = wb.active
        # Получаем заголовки (ID пользователей)
        users = [cell.value for cell in sheet[1][1:] if cell.value]

        # Определяем индексы строк для каждого параметра
        param_indices = {}
        required_params = [
            'SAMPLE_SPREADSHEET_ID',
            'UPDATE_INTERVAL_MINUTES',
            'API_YANDEX_MARKET',
            'BUSINESS_ID_YANDEX_MARKET',
            'YANDEX_MARKET_RANGE',
            'MARKET_NAME',
            'USER_EMAIL',
            'PHONE_NUMBER',
            'PRICE_DECREASE_LOWER',
            'PRICE_DECREASE_UPPER',
            'MARKET_WHITE_LIST'
        ]

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            param_name = row[0].value
            if param_name in required_params:
                param_indices[param_name] = row_idx

        user_configs = []
        for user_id in users:
            user_column = None
            for idx, cell in enumerate(sheet[1]):
                if cell.value == user_id:
                    user_column = idx + 1
                    break

            if user_column:
                params = {}
                for param_name, row_idx in param_indices.items():
                    cell_value = sheet.cell(row=row_idx, column=user_column).value
                    params[param_name.lower()] = cell_value

                config = MarketplaceConfig(
                    user_id=user_id,
                    sample_spreadsheet_id=params.get('sample_spreadsheet_id'),
                    update_interval_minutes=params.get('update_interval_minutes', 5),
                    api_yandex_market=params.get('api_yandex_market'),
                    business_id_yandex_market=params.get('business_id_yandex_market'),
                    yandex_market_range=params.get('yandex_market_range'),
                    market_name=params.get('market_name'),
                    user_email=params.get('user_email'),
                    phone_number=params.get('phone_number'),
                    price_decrease_lower=params.get('price_decrease_lower'),
                    price_decrease_upper=params.get('price_decrease_upper'),
                    market_white_list=params.get('market_white_list')  # Добавлен новый параметр
                )
                user_configs.append(config)

        return user_configs

    except Exception as e:
        logger.error(f"❌ Ошибка при чтении конфигурации из Excel файла: {str(e)}")
        raise


async def process_marketplace_data(config: MarketplaceConfig):
    """Асинхронная функция обработки данных маркетплейсов"""
    user_info = config.get_user_info()
    try:
        async with aiohttp.ClientSession() as session:
            while is_running and not shutdown_event.is_set():
                start_time = datetime.now()
                logger.warning(f"🔄 Начало обработки данных для пользователя {user_info}")

                try:
                    results = []
                    # Обработка данных Яндекс.Маркет
                    if config.has_user_config():
                        logger.info(f"🎁 Обработка данных Яндекс.Маркет для пользователя {user_info}")
                        result = await process_yandex_market_data(session, config)
                        results.append(result)

                    logger.info(f"✅ Завершена обработка данных для пользователя {user_info}")
                    logger.debug(f"Результаты обработки: {json.dumps(results, indent=2)}")

                except Exception as e:
                    logger.error(f"❌ Ошибка при обработке данных для пользователя {user_info}: {str(e)}")

                if shutdown_event.is_set():
                    logger.info(f"🛑 Получен сигнал завершения для пользователя {user_info}")
                    break

                    # Вычисляем время до следующего запуска
                processing_time = (datetime.now() - start_time).total_seconds()
                sleep_time = max(0, math.ceil(config.update_interval_minutes) * 60 - processing_time)

                try:
                    logger.info(
                        f"💤 Пользователь {user_info} - ожидание {sleep_time:.1f} секунд до следующего обновления")
                    await asyncio.wait_for(shutdown_event.wait(), timeout=sleep_time)
                    break  # Если получили событие завершения, прерываем цикл
                except asyncio.TimeoutError:
                    continue  # Если таймаут истек, продолжаем работу

    except asyncio.CancelledError:
        logger.info(f"🛑 Задача для пользователя {user_info} была отменена")
    except Exception as e:
        logger.error(f"❌ Критическая ошибка в обработке данных для пользователя {user_info}: {str(e)}")
    finally:
        logger.info(f"🏁 Задача для пользователя {user_info} завершена")


async def shutdown(signal_name):
    """Корректное завершение всех задач"""
    global is_running
    logger.warning(f"🛑 Получен сигнал {signal_name}. Начинаем корректное завершение...")

    is_running = False
    shutdown_event.set()

    # Ждем завершения текущих задач с таймаутом
    if active_tasks:
        logger.info(f"⏳ Ожидание завершения {len(active_tasks)} активных задач...")
        try:
            await asyncio.wait(active_tasks, timeout=30)  # 30 секунд на завершение
        except Exception as e:
            logger.error(f"❌ Ошибка при ожидании завершения задач: {e}")

            # Принудительная отмена незавершенных задач
        remaining_tasks = {t for t in active_tasks if not t.done()}
        if remaining_tasks:
            logger.warning(f"⚠️ Принудительное завершение {len(remaining_tasks)} задач")
            for task in remaining_tasks:
                task.cancel()

            await asyncio.wait(remaining_tasks)

    logger.warning("✅ Все задачи завершены")


def handle_shutdown(signum, frame):
    """Обработчик сигналов завершения"""
    signal_name = signal.Signals(signum).name
    asyncio.create_task(shutdown(signal_name))


async def main():
    """Основная функция"""
    try:
        # Получаем конфигурации всех пользователей
        user_configs = get_users_config_from_excel('config.xlsx')
        logger.info(f"🚀 Запуск обработки данных для {len(user_configs)} пользователей")

        # Создаем и запускаем задачи для всех пользователей
        for config in user_configs:
            task = asyncio.create_task(
                process_marketplace_data(config),
                name=f"task_{config.user_id}"
            )
            active_tasks.add(task)
            task.add_done_callback(active_tasks.discard)
            logger.info(f"✨ Создана задача для пользователя {config.get_user_info()}")

            # Ждем завершения всех задач или сигнала завершения
        while active_tasks and is_running:
            done, _ = await asyncio.wait(active_tasks, timeout=1)
            for task in done:
                try:
                    await task
                except Exception as e:
                    logger.error(f"❌ Ошибка в задаче: {e}")

    except Exception as e:
        logger.error(f"❌ Ошибка в главной функции: {str(e)}")
    finally:
        if active_tasks:
            await shutdown("FINAL")
        logger.info("🏁 Работа программы завершена")


if __name__ == "__main__":
    # Регистрируем обработчики сигналов для корректного завершения
    signal.signal(signal.SIGINT, handle_shutdown)
    signal.signal(signal.SIGTERM, handle_shutdown)

    # Запускаем асинхронное выполнение
    logger.info("🎯 Запуск программы обработки данных маркетплейсов")
    asyncio.run(main())